import openpyxl 
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from datetime import datetime
import pandas as pd
import win32com.client
from win32com.client import Dispatch

def add_meeting(row) :
    outlook = win32com.client.Dispatch("Outlook.Application") #operate app
    namespace = outlook.GetNamespace("MAPI") #get api

    recipient = namespace.createRecipient("b@dsny.nyc.gov") # input my email
    recipient.Resolve()
    calendar = namespace.GetSharedDefaultFolder(recipient, 9) #get share calender
    appointment = calendar.Items.Add(1) # 1=outlook appointment item ( create events)

    #compose subject
    subject = str(row[0]) + ' - ' + str(row[2])
    startTime = row[5].strftime("%y-%m-%d %H:%M:%S")
    endTime = row[6].strftime("%y-%m-%d %H:%M:%S")
    appointment.Start = startTime #'2023-05-25 10:00' - 6/8/2023  6:00:23 PM
    appointment.end = endTime
    appointment.Subject = subject
    appointment.Categories = "Green Category"
    body = "Change Numnber : \t" + str(row[0])+ "\n"
    body += "Short description : \t" + str(row[2]) + "\n"
    body += "Start Time : \t" + str(row[5]) + "\n"
    body += "End Time : \t" + str(row[6]) + "\n"
    body += "Assigned Group : \t" + str(row[8]) +  "\n"
    body += "Assigned To : \t" + str(row[7]) + "\n"
    body += "Description : \t" + str(row[3]) + "\n"

    appointment.body = body
    appointment.Save()


#Read master change file
book2 = openpyxl.load_workbook("Masterfile.xlsx") 
sheet2 = book2.active 
xrow2 = sheet2.max_row
master_change = []
for x2 in range(2, xrow2 + 1):
    master_change.append(sheet2.cell(row = x2, column = 1).value)

date = datetime.now()
date_str = str(date.month) + str(date.day) + str(date.year)
book2.save('backup/DSNY/master DSNY change backup_' + date_str + '.xlsx')

# Read daily file and find delta
path = "Change2.xlsx"
book = openpyxl.load_workbook(path) 
sheet = book.active 

xrow = sheet.max_row
xcolumn = sheet.max_column

#arr = []
for x in range(2, xrow + 1):
    row = []
    current_number = sheet.cell(row = x, column = 1).value
    if current_number not in master_change:
        for y in range(1, xcolumn - 1): #omitting last two columns - created by  & vreated
            row.append(sheet.cell(row = x, column = y).value)
        #impact = sheet.cell(row = x, column = 9).value
        #print(impact)
        #if 'impact' in impact:
        row.append('Yes')
        row.append('Network')
        row.append('Network')
        row.append('TBD')
        sheet2.append(row)
        add_meeting(row)
    #loops to find the row in the master
    for x2 in range(2, xrow2 + 1):
        #gets the change number at the row
        master_number = sheet2.cell(row = x2, column = 1).value
        #if the change number is the same as the change number in the master
        if(current_number == master_number):
            #get their start & end dates
            current_start = sheet.cell(row = x, column = 6).value
            current_end = sheet.cell(row = x, column = 7).value
            master_start = sheet2.cell(row = x2, column = 6).value
            master_end = sheet2.cell(row = x2, column = 7).value
            yellow = "00FFFF00" # yellow highlight color
            #if both have changed
            if(current_start != master_start and current_end != master_end):
                #change both
                master_start = current_start
                master_end = current_end
                #highlight the change in time in the master
                sheet2.cell(row = x2, column = 6).fill = PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
                sheet2.cell(row = x2, column = 7).fill = PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
            #if only startime changed
            elif(current_start != master_start):
                #change start time
                master_start = current_start
                #highlight change in master
                sheet2.cell(row = x2, column = 6).fill = PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
            #if only the end time changed
            elif(current_end != master_end):
                #change the end time
                master_end = current_end
                #highlight the change in the master
                sheet2.cell(row = x2, column = 7).fill = PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
            #if there are no changes, it does nothing.
        
for rows in sheet2.iter_rows(min_row=1, max_row=None, min_col=None, max_col=None):
    for cell in rows:
       cell.alignment = Alignment(wrapText=True,vertical='top')
df = pd.read_excel('Masterfile.xlsx')
Final_result = df.sort_values('Planned start date')
book2.save("Masterfile.xlsx") 





